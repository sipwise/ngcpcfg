#!/usr/bin/env python3

"""
Generate network.yml files from Excel sheets
============================================

This script takes an Excel file (parsed by openpyxl)
and generates the corresponding ``network.yml`` with
the data contained in the Excel sheets.
"""

import sys
import logging
import argparse
import contextlib

# input formatting
import openpyxl

# output formatting
from ruamel import yaml

from . import objects

DESCRIPTION = 'Generate YAML network configuration from Excel sheets'


@contextlib.contextmanager
def open_or_stdout(src=None, *args, **kwargs):
    """Abstraction of ``open()`` as context manager
    accepting file paths or "-" for stdout::

    >>> with open_or_stdout(filepath) as fd:
    >>>   fd.write(b'data')

    >>> with open_or_stdout("-") as fd:
    >>>   fd.write(b'data')  # writes to sys.stdout
    """
    if src and src != '-':
        fd = open(src, *args, **kwargs)
    else:
        fd = sys.stdout

    try:
        yield fd
    finally:
        if fd is not sys.stdout:
            fd.close()


def convert(src, dst):
    """Convert the XLS network specification to YAML

    :param src:   file to read a binary XLS file from
    :type src:    str or file descriptor
    :param dst:   text file descriptor to write YAML data to
    :type dst:    file descriptor
    """
    # read Excel sheet
    wb = openpyxl.load_workbook(src)

    ws = wb.active  # active work sheet
    msg = "Reading sheet '{}' with {} rows"
    logging.info(msg.format(ws.title, ws.max_row))

    # iterate rows to retrieve all hosts and network connections
    header = {}
    hosts = {}
    vlans = {}
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            for j, cell in enumerate(row):
                if not cell.value or not str(cell.value).strip():
                    continue
                header[j] = str(cell.value).strip().lower()
            continue

        if not row[0].value or not row[1].value:
            continue

        net_data = {}
        net_data['ip_addr'] = row[0].value
        for j, heading in header.items():
            net_data[heading] = row[j].value

        if 'host' not in net_data or not net_data['host']:
            # skip entries without hostname
            continue

        if net_data['host'] not in hosts:
            hosts[net_data['host']] = objects.Host(net_data['host'])

        if net_data['vlan-id'] not in vlans:
            vlan = objects.VirtualLAN(net_data['vlan'], net_data['vlan-id'])
            vlans[net_data['vlan-id']] = vlan

        if net_data['host'].lower() == 'gateway':
            vlans[net_data['vlan-id']].gateway = hosts[net_data['host']]

        net = objects.NetworkConnection(
            ip_addr=net_data['ip_addr'],
            subnetmask=net_data['subnetmask'],
            vlan=vlans[net_data['vlan-id']],
            phy=net_data['phy']
        )

        hosts[net_data['host']].add_net(net)

    # dump data structure YAML to retrieve a ``network.yml`` representation
    sorted_hosts = sorted(hosts.items(), key=lambda h: h[0])
    hostlist = dict((str(hostname), h.yaml()) for hostname, h in sorted_hosts)
    dst.write(yaml.dump({'hosts': hostlist}, default_flow_style=False))
    dst.write('\n')


def main(excelfile, output, loglevel='WARN', output_encoding='utf-8'):
    """Main routine.

    :param excelfile:   filepath to an excelfile
    :type excelfile:    str | file descriptor
    :param loglevel:    logging module log level name
    :type loglevel:     str
    """
    dst = output or '-'

    # configure default logging handler
    logging.basicConfig(level=getattr(logging, loglevel))

    with open_or_stdout(dst, mode='w', encoding='utf-8') as dest:
        convert(excelfile, dest)

    if dst != '-':
        logging.info("Output written to file '{}'".format(dst))


def cli():
    """An entry point for command line interfaces.
    The arguments of sys.argv will be parsed.
    """
    parser = argparse.ArgumentParser(description=DESCRIPTION)
    parser.add_argument('excelfile',
                        help='an Excel 2010 xlsx/xlsm file to read')
    parser.add_argument('-l', '--loglevel', default='WARN',
                        choices='CRITICAL ERROR WARN INFO DEBUG'.split(),
                        help='log level for logging module')
    parser.add_argument('-o', '--output', default='-',
                        help='write output to given file path, not stdout')
    parser.add_argument('--output-encoding', default='utf-8',
                        help='character encoding of the YAML output')
    args = parser.parse_args()

    main(**vars(args))


if __name__ == '__main__':
    cli()
